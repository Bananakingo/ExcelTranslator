import openpyxl
from googletrans import Translator
import xlsxwriter

# Paths
source = "Artikelen Engels Duits.xlsx"
dest = "Map1.xlsx"

# Simple variables
readrow = 1
writerow = 2
translator = Translator()

# 'Read' function adds to list 'readData'
readData = []

# Set sheet object for simplification
sheet_obj = openpyxl.load_workbook(source).active

#  write function
worksheet = xlsxwriter.Workbook(dest).add_worksheet()

while writerow < sheet_obj.max_row:

    # Read 1st column in excel file from path
    cell_obj = sheet_obj.cell(row=writerow, column=2)

    # Writing input to column 3 row by row
    print(cell_obj.value)
    worksheet.write(readrow, 2, translator.translate(cell_obj.value, dest="du", src="nl").text)

    writerow += 1
    readrow += 1

xlsxwriter.Workbook(dest).close()
